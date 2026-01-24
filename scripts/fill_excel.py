
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Define the data based on the analysis
data = [
    {
        "序号": 1,
        "2026 年必须完成事项": "完成全球合伙人后台管理系统的上线",
        "具体项目名称": "全球合伙人系统 (Global Partner System)",
        "核心目标": "支持渠道拓展，提升合作伙伴对接效率与信任度，实现后台管理自动化。",
        "重点任务拆解": "1. 合作伙伴API对接标准文档与落地页发布.\n2. 合伙人后台管理系统开发（权限、订单查看、佣金结算）。\n3. 渠道单量与转化数据看板建设。",
        "完成时限": "2026 Q1",
        "量化KPI指标": "1. 系统上线并稳定运行。\n2. 合作伙伴自助对接率提升。\n3. Q1完成核心功能交付。"
    },
    {
        "序号": 2,
        "2026 年必须完成事项": "协同市场营销部门，建设面向全球客户的高质量网站、APP 及小程序；线上营销多渠道整合。",
        "具体项目名称": "官网 V3 & 营销策略中台 (Online Marketing)",
        "核心目标": "流量获取、线索转化、灵活配置（CMS），打破数据孤岛，实现从数据到行动的闭环。",
        "重点任务拆解": "1. 官网 V3 (CMS动态版) 上线，支持运营自主配置。\n2. 营销策略中台落地 (Data -> Action)，打通Segment/Klaviyo等工具。\n3. SEO & GEO 内容 AI 工作流持续优化。\n4. APP 与 小程序功能迭代适配。",
        "完成时限": "官网 V3: 2026 Q1\n营销中台: 2026 Q2",
        "量化KPI指标": "1. 官网SEO排名与流量提升。\n2. 营销线索转化率提升。\n3. 运营配置效率提升 50%。"
    },
    {
        "序号": 3,
        "2026 年必须完成事项": "于 2026 年上半年完成 JETBAY SOS 产品的核心产品交付。",
        "具体项目名称": "JETBAY SOS 保障体系",
        "核心目标": "建立高可用的紧急救援与保障服务系统，提升客户信任与安全感。",
        "重点任务拆解": "1. SOS 产品核心流程定义与系统开发。\n2. 高可用架构搭建 (SLA 99.95%)。\n3. 应急响应与调度中心功能实现。",
        "完成时限": "2026 H1 (核心交付)",
        "量化KPI指标": "1. 核心产品功能按期上线。\n2. 系统可用性 SLA 99.95%。"
    },
    {
        "序号": 4,
        "2026 年必须完成事项": "完成会员体系与小时卡产品的系统化落地。",
        "具体项目名称": "会员体系与小时卡 (Member & Hourly Card)",
        "核心目标": "提升用户留存，增加预充值现金流，丰富产品矩阵。",
        "重点任务拆解": "1. 会员权益体系配置化开发。\n2. 小时卡产品交易链路打通与UI重构。\n3. 会员与积分系统与 CRM 深度打通。",
        "完成时限": "2026 Q2 (争取提前)",
        "量化KPI指标": "1. 会员系统与小时卡产品功能 100% 上线。\n2. 支持灵活的会员策略配置。"
    },
    {
        "序号": 5,
        "2026 年必须完成事项": "稳定支撑全球包机交易与客户管理系统运行；建立核心经营数据的可视化与分析能力。",
        "具体项目名称": "核心系统重构 (三户中心/订单中心) & BI建设",
        "核心目标": "解决“数据孤岛”，重构老旧架构，提升系统稳定性与扩展性；实现数据驱动决策。",
        "重点任务拆解": "1. 三户中心 (用户/账户/客户) 领域模型重构与数据迁移。\n2. 订单中心核心表结构迁移与服务拆分。\n3. 经营数据可视化看板搭建 (BI)。",
        "完成时限": "2026 Q1 (重构基础)",
        "量化KPI指标": "1. 重构模块数据迁移准确率 100%。\n2. 核心交易链路零故障。\n3. 数据看板覆盖核心经营指标。"
    },
    {
        "序号": 6,
        "2026 年必须完成事项": "提升系统在多语言、多币种、多时区环境下的稳定性；Web3支付。",
        "具体项目名称": "全球化基础设施 & Web3 支付",
        "核心目标": "支持全球业务开展，提升支付自动化率与资金安全。",
        "重点任务拆解": "1. Web3 支付链路全自动化打通 (USDT等)。\n2. 多语言、多时区适配优化。\n3. 资金流安全风控机制升级。",
        "完成时限": "2026 Q2",
        "量化KPI指标": "1. Web3 支付自动化率 100%。\n2. 系统原因导致的资金损失为 0。"
    },
    {
        "序号": 7,
        "2026 年必须完成事项": "持续提升线上业务与客户服务的系统效率；人员效能改变。",
        "具体项目名称": "AI 效能升级与组织转型 (AI-Native)",
        "核心目标": "从“按部就班”转向“AI原生特种部队”，提升研发人效。",
        "重点任务拆解": "1. 落地“双战队”模式 (增长/履约)。\n2. 后端转全栈，利用 AI 生成后台管理页面。\n3. 强制 TDD (测试驱动开发)，自动化测试覆盖。\n4. 建立 AI 辅助的 DevOps 流水线。",
        "完成时限": "2026 全年 (Q1重点落地)",
        "量化KPI指标": "1. 研发交付周期缩短 30%。\n2. 核心模块代码覆盖率达到 80%。\n3. 自动化测试占比大幅提升。"
    }
]

# Create DataFrame
df_new = pd.DataFrame(data)

# File paths
base_dir = "/Users/apus/Desktop/2026年规划/"
filename = "JETBAY各部门2026年度重点工作清单.xlsx"
path = os.path.join(base_dir, filename)

try:
    # 1. Check if file exists to load it, otherwise create new (though user said fill it)
    # We will try to preserve existing style if possible, but pandas writes new sheets by default or overwrites.
    # To append or fill, we can use openpyxl directly or pandas ExcelWriter.
    # Given the requirement is "fill", and likely the file is a template.

    # Let's load with pandas to check if data is there? No, we assume we need to fill it.
    # We will overwrite the '部门名称' sheet or the relevant sheet. 
    # The extraction showed sheet name is '部门名称'. We should probably rename it to '技术产品部' or just fill it.
    # Let's write to the sheet named '技术产品部' if possible, or just overwrite the first sheet if it's a template.
    # The user said "JETBAY各部门...", implies maybe multiple sheets? 
    # But the template has sheet '部门名称'. I will rename it to '技术产品部' and fill data.
    
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_new.to_excel(writer, sheet_name='技术产品部', index=False)
        
    # Re-open to adjust styles
    wb = load_workbook(path)
    if '技术产品部' in wb.sheetnames:
        ws = wb['技术产品部']
        
        # Style formatting
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                cell.border = thin_border
        
        # Adjust column widths
        column_widths = {'A': 5, 'B': 30, 'C': 25, 'D': 30, 'E': 40, 'F': 15, 'G': 30}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    # Remove the default '部门名称' sheet if it's empty and we created a new one, 
    # but 'replace' mode in pandas might have handled it. 
    # If the original sheet was '部门名称' and we wrote to '技术产品部', both might exist.
    # Let's check and clean up.
    if '部门名称' in wb.sheetnames and '技术产品部' in wb.sheetnames:
         # Check if '部门名称' is empty (just header). If so, remove it.
         ws_old = wb['部门名称']
         if ws_old.max_row <= 1:
             del wb['部门名称']

    wb.save(path)
    print(f"Successfully updated {path}")

except Exception as e:
    print(f"Error writing to Excel: {e}")
